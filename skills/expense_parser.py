"""
Parses Pana Studio's expense Excel files into a clean structured dict.
Handles both the historical Expense Tracker format and the 2025 format.
"""

import openpyxl
from datetime import datetime


def _is_date_like(val) -> bool:
    if isinstance(val, datetime):
        return True
    if isinstance(val, str):
        txt = val.strip()
        # Matches patterns like "26th September 2023", "10th Jan 2025", "14th October 2023"
        import re
        return bool(re.match(r'\d+\w*\s+(January|February|March|April|May|June|July|August|September|October|November|December)', txt, re.IGNORECASE))
    return False


def _parse_date(val) -> str:
    if isinstance(val, datetime):
        return val.strftime("%d %B %Y")
    return str(val).strip()


def _parse_overall_sheet(ws) -> list[dict]:
    """Parse the 'Overall Total Yearly IncomeOutc' summary sheet."""
    sessions = []
    for row in ws.iter_rows(values_only=True):
        date_val = row[0] if row else None
        if not _is_date_like(date_val):
            continue
        try:
            sessions.append({
                "date": _parse_date(date_val),
                "shoot_expenses": float(row[1]) if row[1] and row[1] != '=' else 0,
                "income": float(row[2]) if row[2] and row[2] != '=' else 0,
                "employee_salary": float(row[3]) if row[3] and row[3] != '=' else 0,
                "profit_loss": float(row[4]) if row[4] and isinstance(row[4], (int, float)) else 0,
                "status": str(row[5]) if row[5] else "",
                "looks": float(row[9]) if row[9] and isinstance(row[9], (int, float)) else 0,
            })
        except (TypeError, ValueError):
            continue
    return sessions


def _parse_shoot_sheet(ws) -> list[dict]:
    """
    Parse per-shoot expense sheets (Expense Tracker / One Stop Service / Brand).
    These have repeating blocks: [date row, header rows, line items, totals row].
    """
    sessions = []
    current_date = None
    current_session = None
    rows = list(ws.iter_rows(values_only=True))

    for row in rows:
        if not any(cell is not None for cell in row):
            # Empty row — end current session
            if current_session and current_session.get("total_expenses"):
                sessions.append(current_session)
                current_session = None
            continue

        date_val = row[0]
        if _is_date_like(date_val):
            if current_session and current_session.get("total_expenses"):
                sessions.append(current_session)
            current_date = _parse_date(date_val)
            current_session = {
                "date": current_date,
                "shoot_type": "",
                "venue": "",
                "shooting_expenses": {},  # item_name -> amount
                "other_expenses": {},
                "brands": [],
                "total_expenses": 0,
                "total_income": 0,
                "profit_loss": 0,
            }
            continue

        if current_session is None:
            continue

        # Header rows — extract shoot type & venue
        if row[1] in ("Multibrand", "Brand") and current_session["shoot_type"] == "":
            current_session["shoot_type"] = str(row[1])
            current_session["venue"] = str(row[2]) if row[2] else ""
            continue

        # Skip repeated header label rows
        if row[0] == "Brand" and row[3] == "Shooting Expenses":
            continue
        if row[3] in ("(Main)", "(Main) - Full", "(Main) - Half", "(Other)"):
            continue

        # Totals row: col 4 = shooting total, col 7 = other total, col 9 = grand total, col 11 = income, col 13 = P/L
        col4 = row[4] if len(row) > 4 else None
        col7 = row[7] if len(row) > 7 else None
        col9 = row[9] if len(row) > 9 else None
        col11 = row[11] if len(row) > 11 else None
        col13 = row[13] if len(row) > 13 else None

        if isinstance(col4, (int, float)) and isinstance(col9, (int, float)) and col4 > 100:
            current_session["total_expenses"] = float(col9)
            current_session["total_income"] = float(col11) if isinstance(col11, (int, float)) else 0
            current_session["profit_loss"] = float(col13) if isinstance(col13, (int, float)) else 0
            if current_session.get("total_expenses"):
                sessions.append(current_session)
                current_session = None
            continue

        # Brand income line (col 10 = brand name, col 11 = income amount)
        brand_name = row[10] if len(row) > 10 else None
        income_val = row[11] if len(row) > 11 else None
        if brand_name and isinstance(brand_name, str) and isinstance(income_val, (int, float)):
            current_session["brands"].append({"brand": brand_name, "income": float(income_val)})

        # Shooting expense line (col 3 = item, col 4 = amount)
        expense_item = row[3] if len(row) > 3 else None
        expense_amt = row[4] if len(row) > 4 else None
        if expense_item and isinstance(expense_item, str) and isinstance(expense_amt, (int, float)) and expense_amt > 0:
            item_clean = str(expense_item).strip().rstrip("_").strip()
            current_session["shooting_expenses"][item_clean] = (
                current_session["shooting_expenses"].get(item_clean, 0) + float(expense_amt)
            )

        # Other expense line (col 6 = item, col 7 = amount)
        other_item = row[6] if len(row) > 6 else None
        other_amt = row[7] if len(row) > 7 else None
        if other_item and isinstance(other_item, str) and isinstance(other_amt, (int, float)) and other_amt > 0:
            item_clean = str(other_item).strip().strip("-").strip()
            current_session["other_expenses"][item_clean] = (
                current_session["other_expenses"].get(item_clean, 0) + float(other_amt)
            )

    if current_session and current_session.get("total_expenses"):
        sessions.append(current_session)

    return sessions


def _parse_ads_sheet(ws) -> list[dict]:
    """Parse the Ads tracking sheet."""
    ads = []
    for row in ws.iter_rows(values_only=True):
        if not row[0] or not isinstance(row[0], (int, float)):
            continue
        try:
            amount = row[2]
            if not isinstance(amount, (int, float)):
                continue
            ads.append({
                "trc": int(row[0]),
                "start_date": _parse_date(row[1]) if row[1] else "",
                "amount": float(amount),
                "days": float(row[3]) if isinstance(row[3], (int, float)) else 0,
            })
        except (TypeError, ValueError):
            continue
    return ads


def parse_expense_file(file_path: str) -> dict:
    """
    Main entry point. Returns a structured dict with all parsed data.
    """
    wb = openpyxl.load_workbook(file_path, data_only=True)
    result = {
        "file": file_path,
        "sessions": [],
        "summary_sessions": [],
        "ads": [],
        "totals": {},
    }

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        name_lower = sheet_name.lower()

        if "overall" in name_lower or "yearly" in name_lower:
            result["summary_sessions"] = _parse_overall_sheet(ws)
            # Extract grand totals from last row
            for row in ws.iter_rows(values_only=True):
                if isinstance(row[1], (int, float)) and row[1] > 100000:
                    result["totals"] = {
                        "total_shoot_expenses": float(row[1]),
                        "total_income": float(row[2]) if isinstance(row[2], (int, float)) else 0,
                        "total_employee_salary": float(row[3]) if isinstance(row[3], (int, float)) else 0,
                        "net_profit_loss": float(row[4]) if isinstance(row[4], (int, float)) else 0,
                    }

        elif "ads" in name_lower:
            result["ads"] = _parse_ads_sheet(ws)

        elif any(k in name_lower for k in ("expense", "one stop", "brand", "sichang", "tracker")):
            sessions = _parse_shoot_sheet(ws)
            for s in sessions:
                s["sheet"] = sheet_name
            result["sessions"].extend(sessions)

    return result


def format_for_ai(parsed: dict) -> str:
    """
    Convert parsed data into a clean text representation for the AI to analyze.
    """
    lines = []

    # Grand totals (if available from summary sheet)
    if parsed["totals"]:
        t = parsed["totals"]
        lines.append("=== GRAND TOTALS (All Time) ===")
        lines.append(f"Total Shooting Expenses: {t['total_shoot_expenses']:,.0f} THB")
        lines.append(f"Total Income: {t['total_income']:,.0f} THB")
        lines.append(f"Total Employee Salary: {t['total_employee_salary']:,.0f} THB")
        lines.append(f"Net Profit/Loss: {t['net_profit_loss']:,.0f} THB")
        lines.append("")

    # Per-session summary
    if parsed["summary_sessions"]:
        lines.append("=== PER-SESSION SUMMARY ===")
        for s in parsed["summary_sessions"]:
            p = s["profit_loss"]
            sign = "+" if p >= 0 else ""
            lines.append(
                f"{s['date']}: Expenses={s['shoot_expenses']:,.0f} | "
                f"Income={s['income']:,.0f} | "
                f"Salary={s['employee_salary']:,.0f} | "
                f"P/L={sign}{p:,.0f} [{s['status']}]"
            )
        lines.append("")

    # Detailed sessions
    if parsed["sessions"]:
        lines.append("=== DETAILED SHOOT SESSIONS ===")
        for s in parsed["sessions"]:
            lines.append(f"\n--- {s['date']} ({s.get('sheet','')}) | {s['shoot_type']} @ {s['venue']} ---")
            if s["shooting_expenses"]:
                lines.append("  Shooting Expenses:")
                for item, amt in s["shooting_expenses"].items():
                    lines.append(f"    {item}: {amt:,.0f}")
            if s["other_expenses"]:
                lines.append("  Other Expenses:")
                for item, amt in s["other_expenses"].items():
                    lines.append(f"    {item}: {amt:,.0f}")
            lines.append(
                f"  TOTAL: {s['total_expenses']:,.0f} | INCOME: {s['total_income']:,.0f} | "
                f"P/L: {s['profit_loss']:+,.0f}"
            )
            if s["brands"]:
                brand_summary = ", ".join(
                    f"{b['brand']} ({b['income']:,.0f})" for b in s["brands"][:5]
                )
                lines.append(f"  Brands: {brand_summary}")

    # Ads spend
    if parsed["ads"]:
        total_ads = sum(a["amount"] for a in parsed["ads"])
        lines.append(f"\n=== ADS SPEND ===")
        for a in parsed["ads"]:
            lines.append(f"  Campaign #{a['trc']}: {a['amount']:,.0f} THB over {a['days']:.0f} days (from {a['start_date']})")
        lines.append(f"  Total Ads Spent: {total_ads:,.0f} THB")

    return "\n".join(lines)
